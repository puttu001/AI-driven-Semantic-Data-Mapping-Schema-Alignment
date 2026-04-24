"""
File Operations Utilities
Handles saving results to CSV files and MongoDB
"""

import os
import pandas as pd
from typing import List, Dict, Optional
from pathlib import Path
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from config.settings import (
    TOP_K_SEARCH_DISPLAY,
    CSV_TABLE_NAME_COL,
    CSV_TABLE_DESC_COL,
    CSV_COLUMN_NAME_COL,
    CSV_COLUMN_DESC_COL,
    CDM_DATA_DIR,
    MAPPING_DATA_DIR,
    MAPPED_OUTPUT_DIR,
    UNMAPPED_OUTPUT_DIR,
    VALIDATION_OUTPUT_DIR
)


def save_uploaded_file(file_content: bytes, filename: str, file_type: str = 'cdm') -> Path:
    """
    Save uploaded file to artifacts directory.
    
    Args:
        file_content: Binary content of the uploaded file
        filename: Original filename
        file_type: Type of file - 'cdm' or 'mapping'
    
    Returns:
        Path to the saved file
    """
    # Choose directory based on file type 
    target_dir = CDM_DATA_DIR if file_type.lower() == 'cdm' else MAPPING_DATA_DIR
    
    # Add timestamp to filename to avoid conflicts
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = Path(filename).stem
    extension = Path(filename).suffix
    timestamped_filename = f"{base_name}_{timestamp}{extension}"
    
    # Full path
    file_path = target_dir / timestamped_filename
    
    # Save file
    with open(file_path, 'wb') as f:
        f.write(file_content)
    
    print(f"✅ Saved {file_type} file to: {file_path}")
    return file_path


def get_latest_cdm_file() -> Optional[Path]:
    """
    Get the most recently uploaded CDM file.
    
    Returns:
        Path to the latest CDM file or None if not found
    """
    cdm_files = list(CDM_DATA_DIR.glob("*.csv"))
    if not cdm_files:
        return None
    
    # Sort by modification time, return the most recent
    latest_file = max(cdm_files, key=lambda p: p.stat().st_mtime)
    return latest_file


def get_latest_mapping_file() -> Optional[Path]:
    """
    Get the most recently uploaded mapping file.
    
    Returns:
        Path to the latest mapping file or None if not found
    """
    mapping_files = list(MAPPING_DATA_DIR.glob("*.csv"))
    if not mapping_files:
        return None
    
    # Sort by modification time, return the most recent
    latest_file = max(mapping_files, key=lambda p: p.stat().st_mtime)
    return latest_file


def list_input_files() -> Dict[str, List[str]]:
    """
    List all files in inputs directories.
    
    Returns:
        Dictionary with 'cdm_files' and 'mapping_files' lists
    """
    cdm_files = [f.name for f in CDM_DATA_DIR.glob("*.csv")]
    mapping_files = [f.name for f in MAPPING_DATA_DIR.glob("*.csv")]
    
    return {
        "cdm_files": sorted(cdm_files, reverse=True),  # Most recent first
        "mapping_files": sorted(mapping_files, reverse=True)
    }


def cleanup_old_input_files(keep_last_n: int = 5):
    """
    Clean up old inpt files, keeping only the most recent N files.
    
    Args:
        keep_last_n: Number of recent files to keep in each directory
    """
    for directory in [CDM_DATA_DIR, MAPPING_DATA_DIR]:
        files = sorted(directory.glob("*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
        
        # Remove files beyond the keep limit
        for old_file in files[keep_last_n:]:
            try:
                old_file.unlink()
                print(f"🗑️  Removed old artifact: {old_file.name}")
            except Exception as e:
                print(f"⚠️  Failed to remove {old_file.name}: {e}")


def _format_top_candidates(candidates: List[Dict], key: str = 'term') -> str:
    """
    Format top candidates into a readable string.

    Args:
        candidates: List of candidate dictionaries
        key: Key to extract name from (default: 'term')

    Returns:
        Formatted string with candidate names and scores
    """
    if isinstance(candidates, list) and candidates:
        top3 = candidates[:3]
        return '; '.join([f"{c.get(key, 'N/A')} (Score: {c.get('score', 0):.4f})" for c in top3])
    return ''


def _format_new_term_recommendation(recommendation: Optional[Dict]) -> str:
    """
    Format new term recommendation into a readable string for CSV.

    Args:
        recommendation: Recommendation dictionary from LLM

    Returns:
        Formatted string with recommendation details
    """
    if not recommendation or not isinstance(recommendation, dict):
        return ''

    column_name = recommendation.get('recommended_column_name', 'N/A')
    parent = recommendation.get('recommended_parent', 'N/A')
    is_new_parent = recommendation.get('is_new_parent', False)
    confidence = recommendation.get('confidence_score', 0)

    parent_type = "NEW PARENT" if is_new_parent else "EXISTING PARENT"

    return f"{column_name} | Parent: {parent} ({parent_type}) | Confidence: {confidence:.1f}"


def save_results_to_csv(final_mappings: List[Dict], unmapped_columns: List[Dict],
                       output_suffix: str):
    """Save results to CSV files with only specified columns"""

    # Save final mappings - CSV + CDM + LLM details
    if final_mappings:
        df_final = pd.DataFrame(final_mappings)

        # Create output with ONLY these columns
        df_output = pd.DataFrame()

        # CSV columns (source data)
        df_output['Sr. No.'] = range(1, len(df_final) + 1)
        df_output['App Table name'] = df_final.get('csv_table_name', '')
        df_output['App Table Description'] = df_final.get('csv_table_description', '')
        df_output['App Column Name'] = df_final.get('csv_column_name', '')
        df_output['App Column Description'] = df_final.get('csv_column_description', '')

        # CDM Parent Mapping
        df_output['CDM Parent Mapped'] = df_final.get('cdm_parent_name', '')
        df_output['CDM Parent Definition'] = df_final.get('cdm_parent_definition', '')

        # Top 3 Candidates for Parent Mapping
        if 'parent_candidates' in df_final.columns:
            df_output['Top 3 Candidates for Parent Mapping'] = df_final['parent_candidates'].apply(
                lambda x: _format_top_candidates(x if isinstance(x, list) else [], key='parent_name')
            )
        else:
            df_output['Top 3 Candidates for Parent Mapping'] = ''

        # CDM Column Mapping
        df_output['CDM Column Mapped'] = df_final.get('cdm_column_name', '')
        df_output['CDM Column Definiton'] = df_final.get('cdm_column_definition', '')

        # Top 3 Candidates for Column Mapping
        if 'other_candidates' in df_final.columns:
            df_output['Top 3 Candidates'] = df_final['other_candidates'].apply(
                lambda x: _format_top_candidates(x if isinstance(x, list) else [], key='term')
            )
        else:
            df_output['Top 3 Candidates'] = ''

        # Comprehensive Reason - covers all scenarios (accept/reject/modify/skip)
        df_output['Reason'] = df_final.get('comprehensive_reason', df_final.get('llm_reason', ''))

        # New Recommended Column - populated from LLM recommendation if available
        if 'recommended_new_term' in df_final.columns:
            print(f"DEBUG: recommended_new_term column exists")
            print(f"DEBUG: Non-null values: {df_final['recommended_new_term'].notna().sum()}")
            print(f"DEBUG: Sample values (first 3 non-null):")
            non_null_samples = df_final[df_final['recommended_new_term'].notna()]['recommended_new_term'].head(3)
            for idx, val in non_null_samples.items():
                print(f"  Row {idx}: type={type(val)}, value={val}")

            df_output['New Recommended Column'] = df_final['recommended_new_term'].apply(
                lambda x: _format_new_term_recommendation(x)
            )
        else:
            print(f"DEBUG: recommended_new_term column NOT in dataframe columns")
            df_output['New Recommended Column'] = ''

        try:
            final_filename = MAPPED_OUTPUT_DIR / f"Final_CDM_Mappings{output_suffix}.csv"
            df_output.to_csv(final_filename, index=False)
            print(f"✅ Final mappings saved to: {os.path.abspath(final_filename)}")
        except Exception as e:
            print(f"❌ Error saving final mappings: {e}")

    # Save unmapped columns - CSV format with top 3 suggestions
    if unmapped_columns:
        df_unmapped = pd.DataFrame(unmapped_columns)

        # Create output with extended columns including top 3 suggestions
        df_output = pd.DataFrame()
        df_output['Sr. No.'] = range(1, len(df_unmapped) + 1)
        df_output[CSV_TABLE_NAME_COL] = df_unmapped.get('csv_table_name', '')
        df_output[CSV_TABLE_DESC_COL] = df_unmapped.get('csv_table_description', '')
        df_output[CSV_COLUMN_NAME_COL] = df_unmapped.get('csv_column_name', '')
        df_output[CSV_COLUMN_DESC_COL] = df_unmapped.get('csv_column_description', '')

        # Add top 3 CDM column candidates
        if 'other_candidates' in df_unmapped.columns:
            df_output['Top 3 CDM Column Candidates'] = df_unmapped['other_candidates'].apply(
                lambda x: _format_top_candidates(x if isinstance(x, list) else [], key='term')
            )
        else:
            df_output['Top 3 CDM Column Candidates'] = ''

        # Add top 3 CDM parent candidates
        if 'parent_candidates' in df_unmapped.columns:
            df_output['Top 3 CDM Parent Candidates'] = df_unmapped['parent_candidates'].apply(
                lambda x: _format_top_candidates(x if isinstance(x, list) else [], key='parent_name')
            )
        else:
            df_output['Top 3 CDM Parent Candidates'] = ''

        # Add reason for unmapping
        df_output['Reason'] = df_unmapped.get('comprehensive_reason', df_unmapped.get('error', ''))

        try:
            unmapped_filename = UNMAPPED_OUTPUT_DIR / f"Unmapped_Columns{output_suffix}.csv"
            df_output.to_csv(unmapped_filename, index=False)
            print(f"✅ Unmapped columns saved to: {os.path.abspath(unmapped_filename)}")
        except Exception as e:
            print(f"❌ Error saving unmapped columns: {e}")


def _filter_top_candidates(data: List[Dict], max_candidates: int = 3) -> List[Dict]:
    """
    Filter mappings to keep only top N candidates and parent candidates.

    Args:
        data: List of mapping dictionaries
        max_candidates: Maximum number of candidates to keep (default: 3)

    Returns:
        List of mapping dictionaries with filtered candidates
    """
    filtered_data = []

    for item in data:
        # Create a shallow copy to avoid modifying the original
        filtered_item = item.copy()

        # Filter other_candidates (column candidates) to top N
        if 'other_candidates' in filtered_item and isinstance(filtered_item['other_candidates'], list):
            filtered_item['other_candidates'] = filtered_item['other_candidates'][:max_candidates]

        # Filter parent_candidates to top N
        if 'parent_candidates' in filtered_item and isinstance(filtered_item['parent_candidates'], list):
            filtered_item['parent_candidates'] = filtered_item['parent_candidates'][:max_candidates]

        filtered_data.append(filtered_item)

    return filtered_data


def save_results_to_mongodb(
    final_mappings: List[Dict],
    mongodb_uri: str,
    db_name: str = "cdm_mapping",
    mappings_collection: str = "final_mappings",
    execution_metadata: Optional[Dict] = None
) -> Optional[Dict]:
    """
    Save final mappings to MongoDB.
    Only saves top 3 candidates and top 3 parent candidates.

    Args:
        final_mappings: List of final CDM mapping dictionaries
        mongodb_uri: MongoDB connection URI
        db_name: Database name
        mappings_collection: Collection name for final mappings
        execution_metadata: Additional metadata to attach to all records

    Returns:
        Dict with save statistics or None if failed
    """
    from api.api_client import save_mappings_to_mongodb_via_api

    # Filter to keep only top 3 candidates and top 3 parent candidates
    filtered_final_mappings = _filter_top_candidates(final_mappings, max_candidates=3)

    result = save_mappings_to_mongodb_via_api(
        final_mappings=filtered_final_mappings,
        mongodb_uri=mongodb_uri,
        db_name=db_name,
        mappings_collection=mappings_collection,
        execution_metadata=execution_metadata
    )

    return result


def generate_logical_model_excel(mapped_csv_path: Path, output_suffix: str) -> Optional[Path]:
    """
    Generate a formatted Excel logical data model from the mapped CSV file.
    
    Args:
        mapped_csv_path: Path to the final mapped CSV file
        output_suffix: Suffix to use for the output Excel file
    
    Returns:
        Path to the generated Excel file or None if failed
    """
    try:
        # Read the mapped CSV
        df = pd.read_csv(mapped_csv_path)
        
        # Define column rename mapping
        rename_map = {
            "App Table name": "Entity",
            "App Table Description": "Entity Description",
            "App Column Name": "Attribute",
            "App Column Description": "Attribute Description",
            "CDM Parent Mapped": "Mapped CDM Entity",
            "CDM Column Mapped": "Mapped CDM Term",
            "CDM Parent Definition": "Mapped CDM Definition",
            "Reason": "Mapping Decision"
        }
        
        # Select and rename columns
        ldm_view = df[list(rename_map.keys())].rename(columns=rename_map)
        
        # Drop rows where Entity or Attribute are blank
        ldm_view["Entity"] = ldm_view["Entity"].astype(str).str.strip()
        ldm_view["Attribute"] = ldm_view["Attribute"].astype(str).str.strip()
        ldm_view = ldm_view[
            (ldm_view["Entity"] != "") & (ldm_view["Attribute"] != "") &
            (~ldm_view["Entity"].isna()) & (~ldm_view["Attribute"].isna())
        ]
        
        # Sort by Entity, then Attribute
        ldm_view.sort_values(by=["Entity", "Attribute"], inplace=True, ignore_index=True)
        
        # Generate output Excel path
        output_excel = MAPPED_OUTPUT_DIR / f"Logical_Data_Model{output_suffix}.xlsx"
        
        # Write to Excel
        ldm_view.to_excel(output_excel, index=False, sheet_name="Logical_Data_Model")
        
        # Load workbook for formatting
        wb = load_workbook(output_excel)
        ws = wb["Logical_Data_Model"]
        
        # Columns to merge when values are identical vertically
        merge_cols = [
            "Entity", "Entity Description", "Attribute Description",
            "Mapped CDM Entity", "Mapped CDM Term", "Mapped CDM Definition", "Mapping Decision"
        ]
        
        # Create column name → index map
        col_idx_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        
        # Define alignment styles
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        left_no_wrap = Alignment(horizontal="left", vertical="center", wrap_text=False)
        
        def merge_vertical(ws, col_idx, align):
            """Merge consecutive identical cells vertically in one column."""
            start = 2
            prev_value = ws.cell(row=2, column=col_idx).value
            for row in range(3, ws.max_row + 2):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value != prev_value:
                    end = row - 1
                    if end > start:
                        ws.merge_cells(start_row=start, start_column=col_idx,
                                     end_row=end, end_column=col_idx)
                    ws.cell(row=start, column=col_idx).alignment = align
                    start = row
                    prev_value = cell_value
            if start < ws.max_row:
                ws.cell(row=start, column=col_idx).alignment = align
        
        # Merge cells in specified columns
        for name in merge_cols:
            if name not in col_idx_map:
                continue
            idx = col_idx_map[name]
            # Choose alignment type
            if name == "Mapped CDM Definition":
                align = left_align
            elif "Entity" in name:
                align = center_align
            else:
                align = left_no_wrap
            merge_vertical(ws, idx, align)
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 28
        
        # Dynamic row height based on manual line breaks
        for r in range(2, ws.max_row + 1):
            max_lines = 1
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                text = str(cell.value) if cell.value else ""
                line_count = text.count("\n") + 1
                max_lines = max(max_lines, line_count)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            
            base_height = 15
            ws.row_dimensions[r].height = base_height * max_lines
        
        # Auto column width
        for col_idx, col_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            max_len = 0
            for cell in col_cells:
                try:
                    text = str(cell.value) if cell.value else ""
                    text = text.replace("\n", " ").replace("\t", " ")
                    max_len = max(max_len, len(text))
                except Exception:
                    pass
            
            col_name = ws.cell(row=1, column=col_idx).value
            auto_width = min(max_len + 6, 90)
            
            # Custom widths for long text columns
            if col_name == "Entity Description":
                ws.column_dimensions[get_column_letter(col_idx)].width = min(auto_width, 30)
            elif col_name == "Mapped CDM Definition":
                ws.column_dimensions[get_column_letter(col_idx)].width = max(auto_width, 60)
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = auto_width
        
        # Save the workbook
        wb.save(output_excel)
        print(f"✅ Logical Data Model Excel saved to: {os.path.abspath(output_excel)}")
        
        return output_excel
        
    except Exception as e:
        print(f"❌ Error generating Logical Data Model Excel: {e}")
        import traceback
        traceback.print_exc()
        return None
