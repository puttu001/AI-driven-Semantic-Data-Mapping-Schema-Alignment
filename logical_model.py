import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# Step 1: Read your final mapping file
CSV_PATH = r"C:\Users\TransOrg\Desktop\DATAMAPPING2nd\artifacts\mapped\Final_CDM_Mappings_interactive_1771571151.csv"  

OUTPUT_EXCEL = "logical_data_model_1211.xlsx"

df = pd.read_csv(CSV_PATH)

# --- Step 2: Select and rename columns exactly as per prompt ---
rename_map = {
    "App Table name": "Entity",
    "App Table Description": "Entity Description",
    "App Column Name": "Attribute",
    "App Column Description": "Attribute Description",
    "CDM Parent Mapped": "Mapped CDM Entity",
    "CDM Column Mapped": "Mapped CDM Term",
    "CDM Parent Definition": "Mapped CDM Definition",
    "Reason": "Mapping Decision"
}

ldm_view = df[list(rename_map.keys())].rename(columns=rename_map)

# --- Step 3: Drop rows where Entity or Attribute are blank ---
ldm_view["Entity"] = ldm_view["Entity"].astype(str).str.strip()
ldm_view["Attribute"] = ldm_view["Attribute"].astype(str).str.strip()
ldm_view = ldm_view[
    (ldm_view["Entity"] != "") & (ldm_view["Attribute"] != "") &
    (~ldm_view["Entity"].isna()) & (~ldm_view["Attribute"].isna())
]

# --- Step 4: Sort by Entity, then Attribute ---
ldm_view.sort_values(by=["Entity", "Attribute"], inplace=True, ignore_index=True)

# --- Step 5: Write to Excel ---
ldm_view.to_excel(OUTPUT_EXCEL, index=False, sheet_name="Logical_Data_Model")

wb = load_workbook(OUTPUT_EXCEL)
ws = wb["Logical_Data_Model"]

# --- Step 6: Merge adjacent duplicates in 7 columns ---
merge_cols = [
    "Entity", "Entity Description", "Attribute Description",
    "Mapped CDM Entity", "Mapped CDM Term", "Mapped CDM Definition", "Mapping Decision"
]

# Create a column name → index map
col_idx_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

def merge_vertical(ws, col_idx, align):
    """Merge consecutive identical cells vertically in one column."""
    start = 2
    prev_value = ws.cell(row=2, column=col_idx).value
    for row in range(3, ws.max_row + 2):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value != prev_value:
            end = row - 1
            if end > start:
                ws.merge_cells(start_row=start, start_column=col_idx,
                               end_row=end, end_column=col_idx)
            ws.cell(row=start, column=col_idx).alignment = align
            start = row
            prev_value = cell_value
    if start < ws.max_row:
        ws.cell(row=start, column=col_idx).alignment = align

# Define alignment styles
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
left_no_wrap = Alignment(horizontal="left", vertical="center", wrap_text=False)

for name in merge_cols:
    if name not in col_idx_map:
        continue
    idx = col_idx_map[name]
    # Choose alignment type
    if name == "Mapped CDM Definition":
        align = left_align
    elif "Entity" in name:
        align = center_align
    else:
        align = left_no_wrap
    merge_vertical(ws, idx, align)

# === STEP 7: Header formatting ===
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
ws.row_dimensions[1].height = 28

# === STEP 8: No wrap text, dynamic row height based on manual line breaks ===
for r in range(2, ws.max_row + 1):
    max_lines = 1  # at least one line per row
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        text = str(cell.value) if cell.value else ""
        
        # Count manual line breaks in text
        line_count = text.count("\n") + 1
        max_lines = max(max_lines, line_count)

        # No wrap text, center vertically
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    
    # Adjust row height based on number of lines
    base_height = 15   # minimum row height for one line
    ws.row_dimensions[r].height = base_height * max_lines

# === STEP 9: Auto column width (adjusted for no wrap) ===
for col_idx, col_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
    max_len = 0
    for cell in col_cells:
        try:
            text = str(cell.value) if cell.value else ""
            text = text.replace("\n", " ").replace("\t", " ")
            max_len = max(max_len, len(text))
        except Exception:
            pass

    col_name = ws.cell(row=1, column=col_idx).value
    auto_width = min(max_len + 6, 90)

    # Optional: custom widths for long text columns
    if col_name == "Entity Description":
        ws.column_dimensions[get_column_letter(col_idx)].width = min(auto_width, 30)
    elif col_name == "Mapped CDM Definition":
        ws.column_dimensions[get_column_letter(col_idx)].width = max(auto_width, 60)
    else:
        ws.column_dimensions[get_column_letter(col_idx)].width = auto_width


# --- Step 9: Save workbook ---
wb.save(OUTPUT_EXCEL)
print(f"✅ Excel file saved successfully to {OUTPUT_EXCEL}")
