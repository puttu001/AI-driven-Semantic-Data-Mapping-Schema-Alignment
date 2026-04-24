"""
Enhanced Model Excel Writer

Generates a multi-sheet Excel workbook containing:
  Sheet 1: Logical Data Model (entity, attributes, data types, PKs, nullability, CDM mapping)
  Sheet 2: Entity Relationships (source/target entities, cardinality, labels)
  Sheet 3: Physical Data Model (physical names, types, constraints, indexes) — if PDM provided
  Sheet 4: DDL Scripts (database-specific CREATE TABLE statements) — if PDM provided

Formatting follows the same patterns as the existing generate_logical_model_excel()
in utils/file_operations.py (blue headers, merged cells, auto-widths).
"""

from typing import Dict, Optional
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.data_mapping.config.settings import MAPPED_OUTPUT_DIR


# Shared formatting constants (matching existing LDM Excel)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Sheet-specific header colors
SHEET_COLORS = {
    "ldm": "4472C4",       # Blue
    "relationships": "548235",  # Green
    "pdm": "BF8F00",       # Dark Gold
    "ddl": "808080",       # Gray
}


def _apply_header_formatting(ws, fill_color: str = "4472C4") -> None:
    """Apply standard header formatting to the first row."""
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 30


def _apply_auto_column_widths(ws, max_width: int = 90) -> None:
    """Auto-adjust column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            text = str(cell.value) if cell.value else ""
            text = text.replace("\n", " ").replace("\t", " ")
            max_len = max(max_len, len(text))

        col_name = ws.cell(row=1, column=col_idx).value or ""
        auto_width = min(max_len + 4, max_width)

        # Minimum widths for readability
        if "Description" in col_name or "Definition" in col_name:
            auto_width = max(auto_width, 40)
        elif auto_width < 12:
            auto_width = 12

        ws.column_dimensions[get_column_letter(col_idx)].width = auto_width


def _merge_vertical(ws, col_idx: int, align: Alignment) -> None:
    """Merge consecutive identical cells vertically in one column."""
    if ws.max_row < 3:
        return

    start = 2
    prev_value = ws.cell(row=2, column=col_idx).value

    for row in range(3, ws.max_row + 2):
        cell_value = ws.cell(row=row, column=col_idx).value if row <= ws.max_row else None
        if cell_value != prev_value:
            end = row - 1
            if end > start:
                ws.merge_cells(
                    start_row=start, start_column=col_idx,
                    end_row=end, end_column=col_idx
                )
            ws.cell(row=start, column=col_idx).alignment = align
            start = row
            prev_value = cell_value

    if start <= ws.max_row:
        ws.cell(row=start, column=col_idx).alignment = align


def _write_ldm_sheet(wb: Workbook, ldm: Dict) -> None:
    """
    Write Sheet 1: Logical Data Model.

    Columns: Entity, Entity Description, Attribute, Attribute Description,
             Logical Data Type, Primary Key, Nullable,
             Mapped CDM Entity, Mapped CDM Term, Mapped CDM Definition
    """
    ws = wb.active
    ws.title = "Logical Data Model"
    ws.sheet_properties.tabColor = SHEET_COLORS["ldm"]

    # Headers
    headers = [
        "Entity", "Entity Description", "Attribute", "Attribute Description",
        "Logical Data Type", "Primary Key", "Nullable",
        "Mapped CDM Entity", "Mapped CDM Term", "Mapped CDM Definition"
    ]
    ws.append(headers)

    # Data rows
    for entity in ldm.get('entities', []):
        entity_name = entity.get('entity_name', '')
        entity_desc = entity.get('entity_description', '')
        cdm_entity = entity.get('cdm_entity', '')

        for attr in entity.get('attributes', []):
            ws.append([
                entity_name,
                entity_desc,
                attr.get('attribute_name', ''),
                attr.get('attribute_description', ''),
                attr.get('logical_data_type', 'Text'),
                "Yes" if attr.get('is_primary_key') else "No",
                "Yes" if attr.get('is_nullable') else "No",
                cdm_entity,
                attr.get('cdm_term', ''),
                attr.get('cdm_definition', ''),
            ])

    # Formatting
    _apply_header_formatting(ws, SHEET_COLORS["ldm"])

    # Build column index map
    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    # Merge cells for repeated entity info
    center_align = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_name in ["Entity", "Entity Description", "Mapped CDM Entity"]:
        if col_name in col_map:
            _merge_vertical(ws, col_map[col_name], center_align)

    # Wrap text for definition column
    if "Mapped CDM Definition" in col_map:
        def_col = col_map["Mapped CDM Definition"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=def_col).alignment = left_wrap

    _apply_auto_column_widths(ws)

    # Set specific widths for definition columns
    if "Mapped CDM Definition" in col_map:
        ws.column_dimensions[get_column_letter(col_map["Mapped CDM Definition"])].width = 60


def _write_relationships_sheet(wb: Workbook, ldm: Dict) -> None:
    """
    Write Sheet 2: Entity Relationships.

    Columns: Source Entity, Source Attribute, Relationship Type,
             Target Entity, Target Attribute, Relationship Name
    """
    ws = wb.create_sheet("Entity Relationships")
    ws.sheet_properties.tabColor = SHEET_COLORS["relationships"]

    headers = [
        "Source Entity", "Source Attribute", "Relationship Type",
        "Target Entity", "Target Attribute", "Relationship Name"
    ]
    ws.append(headers)

    relationships = ldm.get('relationships', [])

    if not relationships:
        ws.append(["No cross-entity relationships detected", "", "", "", "", ""])
    else:
        for rel in relationships:
            ws.append([
                rel.get('source_entity', ''),
                rel.get('source_attribute', ''),
                rel.get('relationship_type', ''),
                rel.get('target_entity', ''),
                rel.get('target_attribute', ''),
                rel.get('relationship_name', ''),
            ])

    _apply_header_formatting(ws, SHEET_COLORS["relationships"])
    _apply_auto_column_widths(ws)


def _write_pdm_sheet(wb: Workbook, pdm: Dict) -> None:
    """
    Write Sheet 3: Physical Data Model.

    Columns: Physical Table, Physical Column, Data Type,
             Primary Key, Nullable, Foreign Key Reference, Index
    """
    ws = wb.create_sheet("Physical Data Model")
    ws.sheet_properties.tabColor = SHEET_COLORS["pdm"]

    headers = [
        "Physical Table", "Physical Column", "Data Type",
        "Primary Key", "Nullable", "Foreign Key Reference", "Index"
    ]
    ws.append(headers)

    for table in pdm.get('tables', []):
        table_name = table.get('physical_table_name', '')
        # Build index lookup for this table
        indexed_cols = set()
        for idx in table.get('indexes', []):
            for col in idx.get('columns', []):
                indexed_cols.add(col)

        for col in table.get('columns', []):
            col_name = col.get('physical_column_name', '')
            ws.append([
                table_name,
                col_name,
                col.get('physical_data_type', ''),
                "Yes" if col.get('is_primary_key') else "No",
                "Yes" if col.get('is_nullable') else "No",
                col.get('fk_references', '') or '',
                "Yes" if col_name in indexed_cols else "No",
            ])

    _apply_header_formatting(ws, SHEET_COLORS["pdm"])

    # Merge Physical Table column
    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    if "Physical Table" in col_map:
        center_align = Alignment(horizontal="center", vertical="center")
        _merge_vertical(ws, col_map["Physical Table"], center_align)

    _apply_auto_column_widths(ws)


def _write_ddl_sheet(wb: Workbook, ddl: str, dialect: str) -> None:
    """
    Write Sheet 4: DDL Scripts.

    Writes the full DDL as readable content, one statement per block of rows.
    """
    ws = wb.create_sheet(f"DDL Scripts ({dialect})")
    ws.sheet_properties.tabColor = SHEET_COLORS["ddl"]

    ws.append([f"DDL Scripts — {dialect.upper()}"])
    ws.append([""])  # blank row

    # Split DDL into lines and write each one
    mono_font = Font(name="Consolas", size=10)
    code_align = Alignment(horizontal="left", vertical="top", wrap_text=False)

    for line in ddl.split("\n"):
        ws.append([line])

    # Format: make header bold, all content in monospace
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = mono_font
        cell.alignment = code_align

    # Set wide column width for DDL content
    ws.column_dimensions["A"].width = 120


def write_model_excel(
    ldm: Dict,
    pdm: Optional[Dict],
    output_suffix: str,
    output_dir: Optional[Path] = None,
) -> Optional[Path]:
    """
    Write the enhanced LDM/PDM Excel workbook.

    Args:
        ldm: Logical Data Model dict (from generate_logical_data_model).
        pdm: Physical Data Model dict (from generate_physical_data_model).
             If None, only LDM and Relationships sheets are written.
        output_suffix: Suffix for filename (e.g. "_models_1234567890").
        output_dir: Output directory. Defaults to MAPPED_OUTPUT_DIR.

    Returns:
        Path to the generated Excel file, or None if failed.
    """
    if not ldm:
        print("  No LDM provided for Excel generation")
        return None

    try:
        out_dir = output_dir or MAPPED_OUTPUT_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"Logical_Data_Model{output_suffix}.xlsx"

        wb = Workbook()

        # Sheet 1: Logical Data Model
        _write_ldm_sheet(wb, ldm)

        # Sheet 2: Entity Relationships
        _write_relationships_sheet(wb, ldm)

        # Sheet 3 & 4: Physical Data Model + DDL (only if PDM provided)
        if pdm:
            _write_pdm_sheet(wb, pdm)
            ddl = pdm.get('ddl', '')
            dialect = pdm.get('dialect', 'postgresql')
            if ddl:
                _write_ddl_sheet(wb, ddl, dialect)

        wb.save(output_path)
        print(f"  Excel model saved to: {output_path}")

        return output_path

    except Exception as e:
        print(f"  Error generating model Excel: {e}")
        import traceback
        traceback.print_exc()
        return None
